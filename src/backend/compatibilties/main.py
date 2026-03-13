from fastapi import FastAPI, UploadFile, File, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse
from pydantic import BaseModel
from urllib.parse import urlencode
from dotenv import load_dotenv
from typing import Any
import asyncio
import math
import os
import uuid
import time
import json
import unicodedata

import pandas as pd
import openpyxl
import httpx


# =========================
# CARGA DE VARIABLES
# =========================
load_dotenv()

# =========================
# ARCHIVOS / DIRECTORIOS
# =========================
UPLOAD_DIR = "uploads"
TOKENS_FILE = "tokens.json"

os.makedirs(UPLOAD_DIR, exist_ok=True)

# =========================
# CONFIG MERCADO LIBRE
# =========================
ML_CLIENT_ID = os.getenv("ML_CLIENT_ID")
ML_CLIENT_SECRET = os.getenv("ML_CLIENT_SECRET")
ML_REDIRECT_URI = os.getenv("ML_REDIRECT_URI")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")

# Chile
ML_AUTH_URL = "https://auth.mercadolibre.cl/authorization"
ML_TOKEN_URL = "https://api.mercadolibre.com/oauth/token"
ML_ME_URL = "https://api.mercadolibre.com/users/me"
ML_API_BASE = "https://api.mercadolibre.com"
ML_DOMAIN_ID = "MLC-CARS_AND_VANS_FOR_COMPATIBILITIES"
ML_SITE_ID = "MLC"

# =========================
# COLUMNAS / ALIAS
# =========================
COLUMN_ALIASES = {
    "ASOCIACION ML": ["ASOCIACION ML"],
    "MARCA": ["MARCA", "Marca"],
    "MODELO": ["MODELO", "Modelo"],
    "CILINDRADA": ["CILINDRADA", "CILINDRADA ABREVIADA"],
    "TRANSMISION": ["TRANSMISION"],
    "DESDE": ["DESDE", "Desde"],
    "HASTA": ["HASTA", "Hasta"],
}

COMPAT_REQUIRED_LOGICAL_COLUMNS = [
    "ASOCIACION ML",
    "MARCA",
    "MODELO",
    "CILINDRADA",
    "TRANSMISION",
    "DESDE",
    "HASTA",
]


# =========================
# HELPERS TOKENS
# =========================
def load_tokens() -> dict:
    try:
        if not os.path.exists(TOKENS_FILE):
            with open(TOKENS_FILE, "w", encoding="utf-8") as f:
                json.dump({}, f, ensure_ascii=False, indent=2)
            return {}

        with open(TOKENS_FILE, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                return {}
            return json.loads(content)
    except Exception:
        return {}


def save_tokens(tokens: dict) -> None:
    with open(TOKENS_FILE, "w", encoding="utf-8") as f:
        json.dump(tokens, f, ensure_ascii=False, indent=2)


def get_token_data(user_id: int | str):
    return TOKENS_BY_USER.get(str(user_id))


def remove_token_data(user_id: int | str) -> None:
    TOKENS_BY_USER.pop(str(user_id), None)
    save_tokens(TOKENS_BY_USER)


def get_first_user_id() -> str | None:
    if not TOKENS_BY_USER:
        return None
    return next(iter(TOKENS_BY_USER.keys()))


def build_token_payload(token_data: dict, user_id: int | str) -> dict:
    expires_in = int(token_data.get("expires_in", 0))
    token_data["user_id"] = int(user_id)
    token_data["expires_at"] = int(time.time()) + expires_in - 60
    return token_data


def _require_ml_env():
    if not ML_CLIENT_ID:
        raise HTTPException(status_code=500, detail="Falta ML_CLIENT_ID en variables de entorno")
    if not ML_CLIENT_SECRET:
        raise HTTPException(status_code=500, detail="Falta ML_CLIENT_SECRET en variables de entorno")
    if not ML_REDIRECT_URI:
        raise HTTPException(status_code=500, detail="Falta ML_REDIRECT_URI en variables de entorno")

def update_job_progress(job_id: str, progress: int, message: str | None = None):
    job = JOBS.get(job_id)
    if not job:
        return

    job["progress"] = max(0, min(100, int(progress)))
    if message is not None:
        job["message"] = message


async def validate_ml_token(access_token: str) -> bool:
    if not access_token:
        return False

    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(
                ML_ME_URL,
                headers={"Authorization": f"Bearer {access_token}"}
            )
        return r.status_code == 200
    except Exception:
        return False


async def refresh_ml_token_internal(user_id: int | str) -> dict:
    _require_ml_env()

    token_data = get_token_data(user_id)
    if not token_data:
        raise HTTPException(status_code=404, detail="No hay token guardado para ese user_id")

    refresh_token = token_data.get("refresh_token")
    if not refresh_token:
        remove_token_data(user_id)
        raise HTTPException(status_code=400, detail="No hay refresh_token guardado para ese user_id")

    payload = {
        "grant_type": "refresh_token",
        "client_id": ML_CLIENT_ID,
        "client_secret": ML_CLIENT_SECRET,
        "refresh_token": refresh_token,
    }

    headers = {
        "accept": "application/json",
        "content-type": "application/x-www-form-urlencoded",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(ML_TOKEN_URL, data=payload, headers=headers)

    if r.status_code >= 400:
        remove_token_data(user_id)
        raise HTTPException(status_code=r.status_code, detail=r.text)

    new_token_data = r.json()
    new_token_data = build_token_payload(new_token_data, user_id)

    TOKENS_BY_USER[str(user_id)] = new_token_data
    save_tokens(TOKENS_BY_USER)

    return new_token_data


async def get_valid_ml_token(user_id: int | str) -> str:
    token_data = get_token_data(user_id)
    if not token_data:
        raise HTTPException(status_code=404, detail="No hay token guardado para ese user_id")

    access_token = token_data.get("access_token")
    expires_at = int(token_data.get("expires_at", 0))
    now = int(time.time())

    if not access_token or now >= expires_at:
        try:
            token_data = await refresh_ml_token_internal(user_id)
            access_token = token_data.get("access_token")
        except HTTPException:
            remove_token_data(user_id)
            raise

    is_valid = await validate_ml_token(access_token)
    if is_valid:
        return access_token

    try:
        token_data = await refresh_ml_token_internal(user_id)
        access_token = token_data.get("access_token")
    except HTTPException:
        remove_token_data(user_id)
        raise HTTPException(status_code=401, detail="Token inválido y no se pudo refrescar")

    is_valid = await validate_ml_token(access_token)
    if not is_valid:
        remove_token_data(user_id)
        raise HTTPException(status_code=401, detail="Token inválido incluso después del refresh")

    return access_token


# =========================
# HELPERS GENERALES
# =========================
def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def normalize_for_compare(value: Any) -> str:
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("-", " ").replace("_", " ")
    text = " ".join(text.split())
    return text


def normalize_year(value: Any) -> int | None:
    if value is None:
        return None
    try:
        if isinstance(value, float) and math.isnan(value):
            return None
        year = int(float(value))
        if year <= 0:
            return None
        return year
    except Exception:
        return None


def build_years_list(desde: Any, hasta: Any) -> list[int]:
    year_from = normalize_year(desde)
    year_to = normalize_year(hasta)

    if not year_from:
        return []

    # Si HASTA viene vacío, nulo, 0 o menor que DESDE,
    # se procesa solo el año DESDE
    if not year_to or year_to == 0 or year_to < year_from:
        return [year_from]

    return list(range(year_from, year_to + 1))


def normalize_engine(value: Any) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, float) and math.isnan(value):
            return ""
        text = str(value).strip()
        return text
    except Exception:
        return normalize_text(value)


def normalize_transmission(value: Any) -> str:
    text = normalize_for_compare(value)
    mapping = {
        "manual": "Manual",
        "mecanico": "Manual",
        "mecanica": "Manual",
        "mecanico manual": "Manual",
        "mecanica manual": "Manual",
        "automatico": "Automática",
        "automatica": "Automática",
        "auto": "Automática",
        "cvt": "CVT",
    }
    return mapping.get(text, normalize_text(value))


def extract_item_id(value: Any) -> str:
    return normalize_text(value).upper()


def get_row_value(row: dict, logical_name: str) -> Any:
    aliases = COLUMN_ALIASES.get(logical_name, [logical_name])
    for alias in aliases:
        if alias in row:
            return row.get(alias)
    return None


def validate_dataframe_columns(df: pd.DataFrame) -> list[str]:
    missing = []
    cols = set(str(c).strip() for c in df.columns)

    for logical_col in COMPAT_REQUIRED_LOGICAL_COLUMNS:
        aliases = COLUMN_ALIASES.get(logical_col, [logical_col])
        if not any(alias in cols for alias in aliases):
            missing.append(f"{logical_col} (aliases: {aliases})")

    return missing


def debug_print_excel_first15(xlsx_path: str, sheet_name: str = "Hoja1", n_rows: int = 5):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
    print("\n==============================")
    print(f"DEBUG EXCEL: Archivo: {os.path.basename(xlsx_path)}")
    print(f"DEBUG EXCEL: Hoja: {sheet_name}")
    print("DEBUG EXCEL: Primeras 15 columnas (headers):")
    for i, h in enumerate(headers, start=1):
        print(f"  {i}. {h}")

    print(f"\nDEBUG EXCEL: Primeras {n_rows} filas (valores 15 primeras columnas):")
    for r in range(2, 2 + n_rows):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        print(f"FILA {r}: {row_vals}")

    print("==============================\n")


# =========================
# APP
# =========================
app = FastAPI(title="Compatibilidades API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# ESTADO GLOBAL
# =========================
JOBS: dict[str, dict] = {}
TOKENS_BY_USER = load_tokens()


# =========================
# STATUS MERCADO LIBRE
# =========================
@app.get("/ml/status")
async def ml_status():
    user_id = get_first_user_id()
    if not user_id:
        return {"connected": False}

    token_data = get_token_data(user_id)
    if not token_data:
        return {"connected": False}

    try:
        await get_valid_ml_token(user_id)
        token_data = get_token_data(user_id)

        return {
            "connected": True,
            "user_id": user_id,
            "has_refresh_token": bool(token_data.get("refresh_token")),
            "expires_in": token_data.get("expires_in"),
            "expires_at": token_data.get("expires_at"),
        }
    except HTTPException:
        return {"connected": False}


@app.get("/ml/debug-token")
async def ml_debug_token(user_id: int):
    token_data = get_token_data(user_id)
    if not token_data:
        raise HTTPException(status_code=404, detail="No hay token guardado para ese user_id")
    return token_data


@app.get("/ml/access-token")
async def ml_access_token(user_id: int):
    access_token = await get_valid_ml_token(user_id)
    token_data = get_token_data(user_id)

    return {
        "ok": True,
        "user_id": int(user_id),
        "access_token": access_token,
        "expires_at": token_data.get("expires_at"),
        "expires_in": token_data.get("expires_in"),
    }


@app.get("/ml/me")
async def ml_me(user_id: int):
    access_token = await get_valid_ml_token(user_id)

    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(
            ML_ME_URL,
            headers={"Authorization": f"Bearer {access_token}"}
        )

    if r.status_code == 401:
        remove_token_data(user_id)
        raise HTTPException(status_code=401, detail="Sesión inválida. Debes reconectar Mercado Libre.")

    if r.status_code >= 400:
        raise HTTPException(status_code=r.status_code, detail=r.text)

    return r.json()


# =========================
# OAUTH MERCADO LIBRE
# =========================
@app.get("/auth/login")
def ml_auth_login():
    _require_ml_env()

    params = {
        "response_type": "code",
        "client_id": ML_CLIENT_ID,
        "redirect_uri": ML_REDIRECT_URI,
    }

    url = f"{ML_AUTH_URL}?{urlencode(params)}"
    return RedirectResponse(url=url)


@app.get("/auth/callback")
async def ml_auth_callback(code: str = Query(...), state: str | None = None):
    _require_ml_env()

    payload = {
        "grant_type": "authorization_code",
        "client_id": ML_CLIENT_ID,
        "client_secret": ML_CLIENT_SECRET,
        "code": code,
        "redirect_uri": ML_REDIRECT_URI,
    }

    headers = {
        "accept": "application/json",
        "content-type": "application/x-www-form-urlencoded",
    }

    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.post(ML_TOKEN_URL, data=payload, headers=headers)

    if r.status_code >= 400:
        raise HTTPException(status_code=r.status_code, detail=r.text)

    token_data = r.json()
    user_id = token_data.get("user_id")

    if not user_id:
        raise HTTPException(status_code=500, detail="No se recibió user_id desde Mercado Libre")

    token_data = build_token_payload(token_data, user_id)

    TOKENS_BY_USER[str(user_id)] = token_data
    save_tokens(TOKENS_BY_USER)

    return RedirectResponse(url=FRONTEND_URL)


@app.post("/auth/refresh")
async def ml_refresh_token(user_id: int):
    new_token_data = await refresh_ml_token_internal(user_id)

    return {
        "ok": True,
        "user_id": int(user_id),
        "access_token": new_token_data.get("access_token"),
        "refresh_token": new_token_data.get("refresh_token"),
        "expires_in": new_token_data.get("expires_in"),
        "expires_at": new_token_data.get("expires_at"),
        "message": "Token renovado correctamente",
    }


@app.post("/auth/logout")
async def ml_logout(user_id: int):
    token_data = get_token_data(user_id)
    if not token_data:
        return {
            "ok": True,
            "message": "No había sesión local para ese user_id"
        }

    remove_token_data(user_id)
    return {
        "ok": True,
        "message": "Sesión local eliminada correctamente"
    }


# =========================
# MODELOS DE RESPUESTA
# =========================
class JobResponse(BaseModel):
    job_id: str
    status: str
    message: str
    progress: int


# =========================
# HELPERS MERCADO LIBRE API
# =========================
async def ml_request(
    method: str,
    path: str,
    access_token: str,
    json_body: dict | None = None,
    params: dict | None = None,
) -> Any:
    url = f"{ML_API_BASE}{path}"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "application/json",
    }

    if json_body is not None:
        headers["Content-Type"] = "application/json"

    async with httpx.AsyncClient(timeout=60) as client:
        r = await client.request(
            method=method,
            url=url,
            headers=headers,
            json=json_body,
            params=params,
        )

    if r.status_code >= 400:

        print("\n================== ML ERROR ==================")
        print("METHOD:", method)
        print("URL:", url)
        print("PARAMS:", json.dumps(params, ensure_ascii=False, indent=2) if params else None)
        print("JSON BODY:", json.dumps(json_body, ensure_ascii=False, indent=2) if json_body else None)
        print("STATUS CODE:", r.status_code)
        print("RESPONSE TEXT:", r.text)
        print("=============================================\n")
    
        raise HTTPException(
            status_code=r.status_code,
            detail=f"ML API error {r.status_code}: {r.text}"
        )

    try:
        return r.json()
    except Exception:
        return {}


def extract_values_list(data: Any) -> list[dict]:
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]

    if isinstance(data, dict):
        values = data.get("values")
        if isinstance(values, list):
            return [x for x in values if isinstance(x, dict)]

        results = data.get("results")
        if isinstance(results, list):
            return [x for x in results if isinstance(x, dict)]

    return []


async def get_top_values(
    access_token: str,
    attribute_id: str,
    known_attributes: list[dict] | None = None,
    limit: int = 200,
) -> list[dict]:
    body = {"limit": limit}
    if known_attributes:
        body["known_attributes"] = known_attributes

    print("\n========== TOP VALUES REQUEST ==========")
    print("ATTRIBUTE ID:", attribute_id)
    print(json.dumps(body, ensure_ascii=False, indent=2))
    print("========================================\n")

    data = await ml_request(
        "POST",
        f"/catalog_domains/{ML_DOMAIN_ID}/attributes/{attribute_id}/top_values",
        access_token,
        json_body=body,
    )

    return extract_values_list(data)


def pick_value_id_by_name(values: list[dict], wanted_name: str) -> str | None:
    wanted = normalize_for_compare(wanted_name)
    if not wanted:
        return None

    for item in values:
        name = normalize_for_compare(item.get("name"))
        if name == wanted:
            return str(item.get("id"))

    for item in values:
        name = normalize_for_compare(item.get("name"))
        if wanted in name:
            return str(item.get("id"))

    for item in values:
        name = normalize_for_compare(item.get("name"))
        if name and name in wanted:
            return str(item.get("id"))

    return None


async def resolve_brand_id(access_token: str, brand_name: str) -> str | None:
    values = await get_top_values(access_token, "BRAND")
    return pick_value_id_by_name(values, brand_name)


async def resolve_model_id(access_token: str, brand_id: str, model_name: str) -> str | None:
    values = await get_top_values(
        access_token,
        "CAR_AND_VAN_MODEL",
        known_attributes=[
            {"id": "BRAND", "value_id": brand_id}
        ],
    )
    return pick_value_id_by_name(values, model_name)


async def resolve_year_id(access_token: str, brand_id: str, model_id: str, year: int) -> str | None:
    values = await get_top_values(
        access_token,
        "YEAR",
        known_attributes=[
            {"id": "BRAND", "value_id": brand_id},
            {"id": "CAR_AND_VAN_MODEL", "value_id": model_id},
        ],
    )
    return pick_value_id_by_name(values, str(year))


async def resolve_engine_id(
    access_token: str,
    brand_id: str,
    model_id: str,
    year_id: str,
    engine_name: str,
) -> str | None:
    if not engine_name:
        return None

    known_attributes = [
        {"id": "BRAND", "value_id": brand_id},
        {"id": "CAR_AND_VAN_MODEL", "value_id": model_id},
        {"id": "YEAR", "value_id": year_id},
    ]

    values = await get_top_values(
        access_token,
        "CAR_AND_VAN_ENGINE",
        known_attributes=known_attributes,
    )

    engine_id = pick_value_id_by_name(values, engine_name)
    if engine_id:
        return engine_id

    target = normalize_for_compare(engine_name)
    for item in values:
        name = normalize_for_compare(item.get("name"))
        if target and target in name:
            return str(item.get("id"))

    return None


async def resolve_transmission_id(
    access_token: str,
    brand_id: str,
    model_id: str,
    year_id: str,
    transmission_name: str,
) -> str | None:
    if not transmission_name:
        return None

    values = await get_top_values(
        access_token,
        "TRANSMISSION_CONTROL_TYPE",
        known_attributes=[
            {"id": "BRAND", "value_id": brand_id},
            {"id": "CAR_AND_VAN_MODEL", "value_id": model_id},
            {"id": "YEAR", "value_id": year_id},
        ],
    )
    return pick_value_id_by_name(values, transmission_name)


async def search_vehicle_products(
    access_token: str,
    brand_id: str,
    model_id: str,
    year_id: str,
    transmission_id: str | None = None,
    engine_id: str | None = None,
) -> list[dict]:
    known_attributes = [
        {"id": "BRAND", "value_ids": [brand_id]},
        {"id": "CAR_AND_VAN_MODEL", "value_ids": [model_id]},
        {"id": "YEAR", "value_ids": [year_id]},
    ]

    if engine_id:
        known_attributes.append({
            "id": "CAR_AND_VAN_ENGINE",
            "value_ids": [engine_id]
        })

    if transmission_id:
        known_attributes.append({
            "id": "TRANSMISSION_CONTROL_TYPE",
            "value_ids": [transmission_id]
        })

    body = {
        "domain_id": ML_DOMAIN_ID,
        "site_id": ML_SITE_ID,
        "known_attributes": known_attributes,
        "limit": 10,
    }

    print("\n======= PRODUCTS SEARCH REQUEST =======")
    print(json.dumps(body, ensure_ascii=False, indent=2))
    print("======================================\n")

    data = await ml_request(
        "POST",
        "/catalog_compatibilities/products_search/chunks",
        access_token,
        json_body=body,
    )

    return extract_values_list(data)


async def search_vehicle_product_id(
    access_token: str,
    brand_id: str,
    model_id: str,
    year_id: str,
    transmission_id: str | None = None,
    engine_id: str | None = None,
) -> str | None:
    results = await search_vehicle_products(
        access_token=access_token,
        brand_id=brand_id,
        model_id=model_id,
        year_id=year_id,
        transmission_id=transmission_id,
        engine_id=engine_id,
    )
    if not results:
        return None

    first = results[0]
    if not first.get("id"):
        return None

    return str(first.get("id"))


async def get_item_detail(access_token: str, item_id: str) -> dict:
    data = await ml_request(
        "GET",
        f"/items/{item_id}",
        access_token,
    )
    if not isinstance(data, dict):
        raise HTTPException(status_code=500, detail=f"Respuesta inválida de /items/{item_id}")
    return data


async def add_user_product_compatibility(
    access_token: str,
    user_product_id: str,
    category_id: str,
    product_id: str,
    creation_source: str = "DEFAULT",
) -> dict:
    body = {
        "domain_id": ML_DOMAIN_ID,
        "category_id": category_id,
        "products": [
            {
                "id": product_id,
                "creation_source": creation_source,
            }
        ]
    }

    data = await ml_request(
        "POST",
        f"/user-products/{user_product_id}/compatibilities",
        access_token,
        json_body=body,
    )

    if isinstance(data, list):
        return {"raw_response": data}

    return data if isinstance(data, dict) else {"raw_response": data}


# =========================
# LÓGICA DE PROCESAMIENTO
# =========================
async def process_vehicle_row(
    access_token: str,
    row: dict,
    job_id: str | None = None,
    row_index: int = 1,
    total_rows: int = 1,
) -> dict:
    def step_progress(step: int, total_steps: int, text: str):
        if not job_id:
            return

        base = 10 + ((row_index - 1) / max(total_rows, 1)) * 80
        row_weight = 80 / max(total_rows, 1)
        current = base + (step / max(total_steps, 1)) * row_weight

        update_job_progress(
            job_id,
            int(current),
            f"Fila {row_index}/{total_rows} - {text}"
        )

    total_steps = 8

    item_id = extract_item_id(get_row_value(row, "ASOCIACION ML"))
    brand_name = normalize_text(get_row_value(row, "MARCA"))
    model_name = normalize_text(get_row_value(row, "MODELO"))
    engine_name = normalize_engine(get_row_value(row, "CILINDRADA"))
    transmission_name = normalize_transmission(get_row_value(row, "TRANSMISION"))
    years = build_years_list(
        get_row_value(row, "DESDE"),
        get_row_value(row, "HASTA")
    )

    step_progress(1, total_steps, "Validando datos de la fila...")

    if not item_id:
        return {
            "ok": False,
            "reason": "Fila sin ASOCIACION ML"
        }

    if not brand_name or not model_name or not years:
        return {
            "ok": False,
            "item_id": item_id,
            "reason": "Faltan datos mínimos: MARCA / MODELO / DESDE"
        }

    step_progress(2, total_steps, "Consultando detalle del item...")
    item_detail = await get_item_detail(access_token, item_id)
    category_id = item_detail.get("category_id")
    user_product_id = item_detail.get("user_product_id")

    if not category_id:
        return {
            "ok": False,
            "item_id": item_id,
            "reason": "El item no devolvió category_id"
        }

    if not user_product_id:
        return {
            "ok": False,
            "item_id": item_id,
            "reason": "El item no devolvió user_product_id"
        }

    step_progress(3, total_steps, "Resolviendo marca...")
    brand_id = await resolve_brand_id(access_token, brand_name)
    if not brand_id:
        return {
            "ok": False,
            "item_id": item_id,
            "user_product_id": user_product_id,
            "category_id": category_id,
            "reason": f"No se encontró BRAND para '{brand_name}'"
        }

    step_progress(4, total_steps, "Resolviendo modelo...")
    model_id = await resolve_model_id(access_token, brand_id, model_name)
    if not model_id:
        return {
            "ok": False,
            "item_id": item_id,
            "user_product_id": user_product_id,
            "category_id": category_id,
            "reason": f"No se encontró CAR_AND_VAN_MODEL para '{model_name}'"
        }

    results = []
    total_years = len(years)

    for year_idx, year in enumerate(years, start=1):
        year_suffix = f" ({year_idx}/{total_years})" if total_years > 1 else ""

        try:
            step_progress(5, total_steps, f"Resolviendo año {year}{year_suffix}...")
            year_id = await resolve_year_id(access_token, brand_id, model_id, year)

            if not year_id:
                results.append({
                    "ok": False,
                    "item_id": item_id,
                    "user_product_id": user_product_id,
                    "category_id": category_id,
                    "year": year,
                    "brand_id": brand_id,
                    "model_id": model_id,
                    "reason": f"No se encontró YEAR para '{year}'",
                    "brand_name": brand_name,
                    "model_name": model_name,
                })
                continue

            step_progress(6, total_steps, f"Resolviendo motor y transmisión {year_suffix}...")
            engine_id = await resolve_engine_id(
                access_token=access_token,
                brand_id=brand_id,
                model_id=model_id,
                year_id=year_id,
                engine_name=engine_name,
            )

            if engine_name and not engine_id:
                results.append({
                    "ok": False,
                    "item_id": item_id,
                    "user_product_id": user_product_id,
                    "category_id": category_id,
                    "year": year,
                    "brand_id": brand_id,
                    "model_id": model_id,
                    "year_id": year_id,
                    "reason": f"No se encontró CAR_AND_VAN_ENGINE para '{engine_name}'",
                    "brand_name": brand_name,
                    "model_name": model_name,
                })
                continue

            transmission_id = await resolve_transmission_id(
                access_token=access_token,
                brand_id=brand_id,
                model_id=model_id,
                year_id=year_id,
                transmission_name=transmission_name,
            )

            if transmission_name and not transmission_id:
                results.append({
                    "ok": False,
                    "item_id": item_id,
                    "user_product_id": user_product_id,
                    "category_id": category_id,
                    "year": year,
                    "brand_id": brand_id,
                    "model_id": model_id,
                    "year_id": year_id,
                    "engine_id": engine_id,
                    "reason": f"No se encontró TRANSMISSION_CONTROL_TYPE para '{transmission_name}'",
                    "brand_name": brand_name,
                    "model_name": model_name,
                })
                continue

            step_progress(7, total_steps, f"Buscando producto compatible {year_suffix}...")
            product_id = await search_vehicle_product_id(
                access_token=access_token,
                brand_id=brand_id,
                model_id=model_id,
                year_id=year_id,
                transmission_id=transmission_id,
                engine_id=engine_id,
            )

            if not product_id:
                results.append({
                    "ok": False,
                    "item_id": item_id,
                    "user_product_id": user_product_id,
                    "category_id": category_id,
                    "year": year,
                    "brand_id": brand_id,
                    "model_id": model_id,
                    "year_id": year_id,
                    "engine_id": engine_id,
                    "transmission_id": transmission_id,
                    "reason": "No se encontró product_id en products_search/chunks",
                    "brand_name": brand_name,
                    "model_name": model_name,
                })
                continue

            step_progress(8, total_steps, f"Grabando compatibilidad para año {year}{year_suffix}...")
            ml_response = await add_user_product_compatibility(
                access_token=access_token,
                user_product_id=str(user_product_id),
                category_id=str(category_id),
                product_id=product_id,
                creation_source="DEFAULT",
            )

            results.append({
                "ok": True,
                "item_id": item_id,
                "user_product_id": user_product_id,
                "category_id": category_id,
                "year": year,
                "product_id": product_id,
                "brand_id": brand_id,
                "model_id": model_id,
                "year_id": year_id,
                "engine_id": engine_id,
                "transmission_id": transmission_id,
                "brand_name": brand_name,
                "model_name": model_name,
                "request_body": {
                    "domain_id": ML_DOMAIN_ID,
                    "category_id": category_id,
                    "products": [
                        {
                            "id": product_id,
                            "creation_source": "DEFAULT"
                        }
                    ]
                },
                "ml_response": ml_response,
            })

        except HTTPException as e:
            results.append({
                "ok": False,
                "item_id": item_id,
                "user_product_id": user_product_id,
                "category_id": category_id,
                "year": year,
                "reason": f"HTTPException procesando año {year}: {e.detail}",
                "brand_name": brand_name,
                "model_name": model_name,
            })
            continue

        except Exception as e:
            results.append({
                "ok": False,
                "item_id": item_id,
                "user_product_id": user_product_id,
                "category_id": category_id,
                "year": year,
                "reason": f"Exception procesando año {year}: {str(e)}",
                "brand_name": brand_name,
                "model_name": model_name,
            })
            continue

    success_results = [r for r in results if r.get("ok")]
    error_results = [r for r in results if not r.get("ok")]

    success_results = [r for r in results if r.get("ok")]
    error_results = [r for r in results if not r.get("ok")]

    return {
        "ok": len(success_results) > 0,
        "item_id": item_id,
        "brand_name": brand_name,
        "model_name": model_name,
        "engine_name": engine_name,
        "transmission_name": transmission_name,
        "user_product_id": user_product_id,
        "category_id": category_id,
        "years_requested": years,
        "years_processed": len(results),
        "success_count": len(success_results),
        "error_count": len(error_results),
        "results": results,
    }

async def process_excel_job_async(job_id: str, user_id: int | str):
    job = JOBS[job_id]
    xlsx_path = job["xlsx_path"]

    job["status"] = "processing"
    job["message"] = "Preparando procesamiento..."
    job["progress"] = 0

    try:
        update_job_progress(job_id, 2, "Validando sesión con Mercado Libre...")
        access_token = await get_valid_ml_token(user_id)

        update_job_progress(job_id, 5, "Leyendo archivo Excel...")
        df = pd.read_excel(xlsx_path, sheet_name="Hoja1", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]

        update_job_progress(job_id, 8, "Validando columnas requeridas...")
        missing = validate_dataframe_columns(df)
        if missing:
            job["status"] = "error"
            job["message"] = f"Faltan columnas requeridas: {missing}"
            job["progress"] = 0
            return

        total_rows = len(df.index)
        if total_rows <= 0:
            job["status"] = "error"
            job["message"] = "El Excel no tiene filas."
            job["progress"] = 0
            return

        update_job_progress(
            job_id,
            10,
            f"Comenzando procesamiento de {total_rows} fila(s)..."
        )

        output_rows = []

        # Resumen por fila Excel
        rows_ok = 0
        rows_error = 0

        # Resumen por compatibilidad real
        compatibilities_total = 0
        compatibilities_ok = 0
        compatibilities_error = 0

        for idx, (_, row) in enumerate(df.iterrows(), start=1):
            row_data = row.to_dict()

            update_job_progress(
                job_id,
                max(10, int(((idx - 1) / max(total_rows, 1)) * 90)),
                f"Preparando fila {idx}/{total_rows}..."
            )

            try:
                result = await process_vehicle_row(
                    access_token=access_token,
                    row=row_data,
                    job_id=job_id,
                    row_index=idx,
                    total_rows=total_rows,
                )
            except HTTPException as e:
                result = {
                    "ok": False,
                    "item_id": extract_item_id(get_row_value(row_data, "ASOCIACION ML")),
                    "brand_name": normalize_text(get_row_value(row_data, "MARCA")),
                    "model_name": normalize_text(get_row_value(row_data, "MODELO")),
                    "engine_name": normalize_engine(get_row_value(row_data, "CILINDRADA")),
                    "transmission_name": normalize_transmission(get_row_value(row_data, "TRANSMISION")),
                    "reason": f"HTTPException: {e.detail}",
                    "results": []
                }
            except Exception as e:
                result = {
                    "ok": False,
                    "item_id": extract_item_id(get_row_value(row_data, "ASOCIACION ML")),
                    "brand_name": normalize_text(get_row_value(row_data, "MARCA")),
                    "model_name": normalize_text(get_row_value(row_data, "MODELO")),
                    "engine_name": normalize_engine(get_row_value(row_data, "CILINDRADA")),
                    "transmission_name": normalize_transmission(get_row_value(row_data, "TRANSMISION")),
                    "reason": f"Exception: {str(e)}",
                    "results": []
                }

            output_rows.append(result)

            # Conteo por fila
            if result.get("ok"):
                rows_ok += 1
            else:
                rows_error += 1

            # Conteo por compatibilidad real
            detail_results = result.get("results", [])
            if isinstance(detail_results, list) and detail_results:
                compatibilities_total += len(detail_results)
                compatibilities_ok += sum(1 for d in detail_results if d.get("ok"))
                compatibilities_error += sum(1 for d in detail_results if not d.get("ok"))
            else:
                # Si no hubo detalle, tratamos la fila fallida como 1 error lógico
                if not result.get("ok"):
                    compatibilities_total += 1
                    compatibilities_error += 1

            progress_after_row = 10 + int((idx / max(total_rows, 1)) * 80)
            update_job_progress(
                job_id,
                progress_after_row,
                (
                    f"Procesadas {idx}/{total_rows} fila(s) | "
                    f"filas ok={rows_ok} | filas error={rows_error} | "
                    f"compat ok={compatibilities_ok} | compat error={compatibilities_error}"
                )
            )

        update_job_progress(job_id, 95, "Guardando resultado final...")

        result_path = os.path.join(UPLOAD_DIR, f"{job_id}_resultado.json")
        with open(result_path, "w", encoding="utf-8") as f:
            json.dump(output_rows, f, ensure_ascii=False, indent=2)

        job["status"] = "success"
        job["message"] = (
            f"Proceso finalizado. "
            f"Filas OK={rows_ok}, filas error={rows_error}, "
            f"compatibilidades OK={compatibilities_ok}, compatibilidades error={compatibilities_error}"
        )
        job["progress"] = 100
        job["result_path"] = result_path
        job["summary"] = {
            # Resumen por filas del Excel
            "processed_rows": total_rows,
            "success_count": rows_ok,
            "error_count": rows_error,

            # Resumen por compatibilidades reales
            "compatibilities_total": compatibilities_total,
            "compatibilities_ok": compatibilities_ok,
            "compatibilities_error": compatibilities_error,

            # Métricas extra útiles
            "items_processed": len(output_rows),
            "items_with_success": sum(1 for r in output_rows if r.get("ok")),
            "items_with_error": sum(1 for r in output_rows if not r.get("ok")),
        }

    except Exception as e:
        job["status"] = "error"
        job["message"] = f"Error procesando Excel: {str(e)}"
        job["progress"] = 0


def process_excel_job(job_id: str, user_id: int | str):
    asyncio.run(process_excel_job_async(job_id, user_id))


# =========================
# ENDPOINTS IMPORT / JOBS
# =========================
@app.post("/imports-excel", response_model=JobResponse)
async def upload_excel(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    job_id = str(uuid.uuid4())
    xlsx_path = os.path.join(UPLOAD_DIR, f"{job_id}.xlsx")

    content = await file.read()
    with open(xlsx_path, "wb") as f:
        f.write(content)

    try:
        debug_print_excel_first15(xlsx_path, sheet_name="Hoja1", n_rows=5)
    except Exception as e:
        print("WARNING: no se pudo imprimir debug del Excel:", str(e))

    try:
        df = pd.read_excel(xlsx_path, sheet_name="Hoja1", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        missing = validate_dataframe_columns(df)
        if missing:
            raise ValueError(f"Faltan columnas requeridas: {missing}")
    except Exception as e:
        JOBS[job_id] = {
            "status": "error",
            "message": f"Error leyendo Excel: {str(e)}",
            "progress": 0,
            "xlsx_path": xlsx_path,
            "created_at": int(time.time()),
        }
        return JobResponse(
            job_id=job_id,
            status="error",
            message=JOBS[job_id]["message"],
            progress=0,
        )

    JOBS[job_id] = {
        "status": "ready",
        "message": "Excel subido correctamente. Listo para procesar compatibilidades.",
        "progress": 0,
        "xlsx_path": xlsx_path,
        "created_at": int(time.time()),
    }

    return JobResponse(
        job_id=job_id,
        status="ready",
        message=JOBS[job_id]["message"],
        progress=0,
    )


@app.post("/imports/{job_id}/start", response_model=JobResponse)
async def start_processing(job_id: str, background_tasks: BackgroundTasks):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no existe")

    if job["status"] in ("processing", "success"):
        return JobResponse(
            job_id=job_id,
            status=job["status"],
            message=job["message"],
            progress=job["progress"],
        )

    if job["status"] == "error":
        return JobResponse(
            job_id=job_id,
            status="error",
            message=job["message"],
            progress=job["progress"],
        )

    if not job.get("xlsx_path"):
        return JobResponse(
            job_id=job_id,
            status="error",
            message="Excel no disponible para procesar.",
            progress=0,
        )

    user_id = get_first_user_id()
    if not user_id:
        raise HTTPException(status_code=401, detail="No hay cuenta de Mercado Libre conectada")

    background_tasks.add_task(process_excel_job, job_id, user_id)
    job["status"] = "processing"
    job["message"] = "Procesamiento en cola..."
    job["progress"] = 0

    return JobResponse(
        job_id=job_id,
        status="processing",
        message=job["message"],
        progress=0,
    )


@app.get("/imports/{job_id}", response_model=JobResponse)
async def get_job(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no existe")

    return JobResponse(
        job_id=job_id,
        status=job["status"],
        message=job["message"],
        progress=job["progress"],
    )


@app.get("/imports/{job_id}/result")
async def get_job_result(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job no existe")

    if job.get("status") != "success":
        raise HTTPException(status_code=400, detail="El job aún no finaliza correctamente")

    result_path = job.get("result_path")
    if not result_path or not os.path.exists(result_path):
        raise HTTPException(status_code=404, detail="No se encontró archivo de resultado")

    with open(result_path, "r", encoding="utf-8") as f:
        result_data = json.load(f)

    return {
        "ok": True,
        "job_id": job_id,
        "summary": job.get("summary", {}),
        "results": result_data,
    }


# =========================
# ENDPOINTS DE PRUEBA / DEBUG
# =========================
@app.post("/compatibilities/test-one-row")
async def compatibilities_test_one_row(user_id: int):
    access_token = await get_valid_ml_token(user_id)

    latest_job = None
    for _, job in JOBS.items():
        if job.get("xlsx_path"):
            latest_job = job

    if not latest_job:
        raise HTTPException(status_code=404, detail="No hay Excel cargado")

    xlsx_path = latest_job["xlsx_path"]
    df = pd.read_excel(xlsx_path, sheet_name="Hoja1", engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    missing = validate_dataframe_columns(df)
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas: {missing}")

    if len(df.index) == 0:
        raise HTTPException(status_code=400, detail="El Excel no tiene filas")

    first_row = df.iloc[0].to_dict()
    result = await process_vehicle_row(access_token, first_row)

    return {
        "ok": True,
        "input_row": first_row,
        "result": result,
    }


@app.get("/compatibilities/item-debug")
async def compatibilities_item_debug(user_id: int, item_id: str):
    access_token = await get_valid_ml_token(user_id)
    item_detail = await get_item_detail(access_token, item_id)

    return {
        "ok": True,
        "item_id": item_id,
        "category_id": item_detail.get("category_id"),
        "user_product_id": item_detail.get("user_product_id"),
        "full_item": item_detail,
    }